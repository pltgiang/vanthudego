"""Test xuất Excel báo cáo mua hàng (khớp form thumua1 sheet 12–16).

Kiểm builder trên dữ liệu tổng hợp: đúng sheet, header, ô số, định dạng %, dòng TỔNG CỘNG,
và tô đỏ tỷ lệ > 30% (NCC / Bộ phận)."""
import openpyxl
import pytest

from app.modules.report.excel import build_report_workbook

MONTHS = [{"key": f"2026-{m:02d}", "label": f"{m:02d}/2026"} for m in range(1, 13)]

DATA = {
    "months": MONTHS,
    "nspt": [
        {"key": "Mr Tiên", "orders": 4, "late": 3, "ontime": 0, "early": 1, "rate": 75.0,
         "m": {"2026-04": {"orders": 4, "late": 3, "ontime": 0, "early": 1, "rate": 75.0}}},
    ],
    "item_group": [
        {"key": "Thùng", "trans": 37, "cost": 403943400,
         "m": {"2026-06": {"trans": 34, "cost": 389163600}, "2026-03": {"trans": 3, "cost": 14779800}}},
    ],
    "supplier": [
        {"key": "Đông Tây", "trans": 16, "late": 3, "rate": 18.75,
         "m": {"2026-06": {"trans": 13, "late": 1, "rate": 7.69}}},
        {"key": "Xấu NCC", "trans": 10, "late": 6, "rate": 60.0,
         "m": {"2026-06": {"trans": 10, "late": 6, "rate": 60.0}}},
    ],
    "department": [
        {"key": "SX - Ms Ly", "orders": 7, "urgent": 1, "rate": 14.29,
         "m": {"2026-06": {"orders": 4, "urgent": 1, "rate": 25.0}}},
    ],
    "shipping": [
        {"key": "Sang Giàu", "freq": 5, "order_value": 1000, "ship_cost": 50, "rate": 5.0,
         "m": {"2026-04": {"freq": 5, "order_value": 1000, "ship_cost": 50, "rate": 5.0}}},
    ],
    "shipping_detail": [
        {"carrier": "Sang Giàu", "month": "04/2026", "product_code": "P1", "misa_code": "M1",
         "invoice_no": "HD1", "received_date": "2026-04-10", "qty_order": 10, "qty_received": 10,
         "order_amount": 1000, "ship_amount": 50, "rate": 5.0},
    ],
}


def _wb(sheet):
    return openpyxl.load_workbook(build_report_workbook(DATA, sheet))


def test_all_has_five_sheets_in_order():
    wb = _wb("all")
    assert wb.sheetnames == [
        "12. BÁO CÁO NSPT", "13. BÁO CÁO PHÂN LOẠI VTBBNL", "14. BÁO CÁO NCC",
        "15. BÁO CÁO BỘ PHẬN", "16. BÁO CÁO CHI PHÍ VẬN CHUYỂN",
    ]


@pytest.mark.parametrize("sheet,name", [
    ("nspt", "12. BÁO CÁO NSPT"), ("item_group", "13. BÁO CÁO PHÂN LOẠI VTBBNL"),
    ("supplier", "14. BÁO CÁO NCC"), ("department", "15. BÁO CÁO BỘ PHẬN"),
    ("shipping", "16. BÁO CÁO CHI PHÍ VẬN CHUYỂN"),
])
def test_single_sheet(sheet, name):
    wb = _wb(sheet)
    assert wb.sheetnames == [name]


def test_invalid_sheet_raises():
    with pytest.raises(ValueError):
        build_report_workbook(DATA, "khong_ton_tai")


def test_nspt_headers_and_values():
    ws = _wb("nspt")["12. BÁO CÁO NSPT"]
    assert ws["A5"].value == "STT" and ws["B5"].value == "NSPT" and ws["C5"].value == "THÁNG & NĂM"
    assert [ws.cell(row=7, column=c).value for c in range(3, 8)] == \
        ["Tổng số đơn", "Tổng đơn bị trễ", "Tổng đơn đúng hẹn", "Tổng đơn giao sớm", "Tỷ lệ trễ đơn"]
    assert ws["B8"].value == "TỔNG CỘNG"
    # Mr Tiên tháng 4 (tháng thứ 4): cột bắt đầu C(3) + 3*5 = 18
    c0 = 3 + 3 * 5
    assert ws.cell(row=9, column=c0).value == 4        # orders
    assert ws.cell(row=9, column=c0 + 1).value == 3    # late
    rate = ws.cell(row=9, column=c0 + 4)
    assert rate.value == pytest.approx(0.75) and rate.number_format == "0.0%"


def test_nspt_year_totals():
    ws = _wb("nspt")["12. BÁO CÁO NSPT"]
    yc = 3 + 12 * 5   # sau 12 tháng × 5 cột con
    assert ws.cell(row=9, column=yc).value == 4        # orders năm
    assert ws.cell(row=9, column=yc + 1).value == 3    # late năm
    assert ws.cell(row=9, column=yc + 2).value == pytest.approx(0.75)


def test_supplier_warn_fill_over_30_percent():
    ws = _wb("supplier")["14. BÁO CÁO NCC"]
    yc_rate = 3 + 12 * 3 + 2   # cột tỷ lệ năm trong block tổng
    fills = {}
    for r in range(9, 12):
        name = ws.cell(row=r, column=2).value
        if name in ("Xấu NCC", "Đông Tây"):
            fills[name] = ws.cell(row=r, column=yc_rate).fill.fgColor.rgb
    assert fills["Xấu NCC"] == "00FFC7CE"      # 60% > 30% -> tô đỏ
    assert fills["Đông Tây"] != "00FFC7CE"     # 18.75% -> không tô


def test_shipping_two_tables():
    ws = _wb("shipping")["16. BÁO CÁO CHI PHÍ VẬN CHUYỂN"]
    assert [ws.cell(row=4, column=c).value for c in range(1, 7)] == \
        ["ĐƠN VỊ VẬN CHUYỂN", "THÁNG", "TẦN SUẤT PHÁT SINH",
         "TỔNG GIÁ TRỊ ĐƠN HÀNG", "TỔNG CHI PHÍ VẬN CHUYỂN", "TỶ LỆ"]
    # bảng 1 có dòng dữ liệu chành
    assert ws.cell(row=5, column=1).value == "Sang Giàu"
    # bảng 2 (chi tiết) xuất hiện phía dưới với đủ 11 cột header
    found = False
    for r in range(6, 20):
        if ws.cell(row=r, column=1).value == "Đơn vị vận chuyển":
            assert ws.cell(row=r, column=11).value == "TỶ LỆ"
            found = True
            break
    assert found
