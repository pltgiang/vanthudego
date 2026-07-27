"""Dựng 1 sheet ma trận: rows = đối tượng, cols = 12 tháng × cột con + block Tổng năm.

Dùng chung cho 4 form: NSPT (12), Phân loại VTBB/NL (13), NCC (14), Bộ phận (15).
Mỗi form chỉ khác: nhãn, danh sách cột con/tháng, cột tổng năm, có tô đỏ >30% hay không."""
from openpyxl.utils import get_column_letter

from . import styles as S


def _set(ws, r, c, value, *, numeric=True, total=False, warn=False, fmt=None):
    cell = ws.cell(row=r, column=c, value=value)
    S.style_body_cell(cell, numeric=numeric, total=total, warn=warn)
    if fmt and value is not None:
        cell.number_format = fmt
    return cell


def build_pivot_sheet(ws, *, desc, title, entity_label, months, rows,
                      subcols, year_cols, warn_col_key=None):
    """
    subcols   : list (header, key, fmt)  — cột con lặp cho mỗi tháng
    year_cols : list (header, key, fmt)  — block "Tổng cộng năm" cuối bảng
    warn_col_key: key trong year_cols sẽ tô đỏ khi row['warn'] = True (NCC/Bộ phận)
    """
    nsub = len(subcols)
    nmonth = len(months)
    year = months[0]["key"][:4] if months else ""

    # ----- Tiêu đề -----
    ws.cell(row=1, column=2, value=desc).font = S.FONT_DESC
    tc = ws.cell(row=3, column=3, value=title)
    tc.font = S.FONT_TITLE
    if warn_col_key:
        ws.cell(row=4, column=3, value="Cảnh báo: tỷ lệ trên 30% được tô màu đỏ").font = S.FONT_WARN_NOTE

    HR1, HR2, HR3 = 5, 6, 7       # 3 tầng header
    DATA0 = 8                     # dòng TỔNG CỘNG
    col_month0 = 3                # cột C bắt đầu các tháng

    # ----- Header cố định: STT / entity / 'THÁNG & NĂM' -----
    a = ws.cell(row=HR1, column=1, value="STT"); S.style_header_cell(a)
    ws.merge_cells(start_row=HR1, start_column=1, end_row=HR3, end_column=1)
    b = ws.cell(row=HR1, column=2, value=entity_label); S.style_header_cell(b)
    ws.merge_cells(start_row=HR1, start_column=2, end_row=HR3, end_column=2)
    last_month_col = col_month0 + nmonth * nsub - 1
    tn = ws.cell(row=HR1, column=col_month0, value="THÁNG & NĂM"); S.style_header_cell(tn)
    ws.merge_cells(start_row=HR1, start_column=col_month0, end_row=HR1, end_column=last_month_col)

    # ----- Header tháng (tầng 2) + cột con (tầng 3) -----
    for mi, mo in enumerate(months):
        c0 = col_month0 + mi * nsub
        mc = ws.cell(row=HR2, column=c0, value=mo["label"]); S.style_header_cell(mc)
        if nsub > 1:
            ws.merge_cells(start_row=HR2, start_column=c0, end_row=HR2, end_column=c0 + nsub - 1)
        for si, (hdr, _k, _f) in enumerate(subcols):
            S.style_header_cell(ws.cell(row=HR3, column=c0 + si, value=hdr))

    # ----- Block Tổng năm -----
    year_col0 = last_month_col + 1
    for yi, (hdr, _k, _f) in enumerate(year_cols):
        yc = ws.cell(row=HR1, column=year_col0 + yi, value=f"{hdr}\nNĂM {year}")
        S.style_header_cell(yc)
        ws.merge_cells(start_row=HR1, start_column=year_col0 + yi, end_row=HR3, end_column=year_col0 + yi)

    mkey = [m["key"] for m in months]

    def write_row(r, label, stt, mfn, yobj, *, total=False, warn=False):
        _set(ws, r, 1, stt, numeric=False, total=total)
        _set(ws, r, 2, label, numeric=False, total=total)
        for mi, mk in enumerate(mkey):
            cell = mfn(mk)
            c0 = col_month0 + mi * nsub
            for si, (_h, key, fmt) in enumerate(subcols):
                v = cell.get(key, 0) if cell else 0
                if fmt == S.FMT_PCT:
                    v = (v or 0) / 100.0
                _set(ws, r, c0 + si, v, total=total, fmt=fmt)
        for yi, (_h, key, fmt) in enumerate(year_cols):
            v = (yobj.get(key, 0) if yobj else 0)
            w = warn and (key == warn_col_key)
            if fmt == S.FMT_PCT:
                v = (v or 0) / 100.0
            _set(ws, r, year_col0 + yi, v, total=total, warn=w, fmt=fmt)

    # ----- Dòng TỔNG CỘNG (gom toàn bộ đối tượng) -----
    grand_m = {mk: {} for mk in mkey}
    grand_y = {}
    numeric_keys = {k for _h, k, f in subcols if f != S.FMT_PCT} | \
                   {k for _h, k, f in year_cols if f != S.FMT_PCT}
    for row in rows:
        for mk in mkey:
            cell = row.get("m", {}).get(mk)
            if cell:
                for k in numeric_keys:
                    if k in cell:
                        grand_m[mk][k] = grand_m[mk].get(k, 0) + cell[k]
        for k in numeric_keys:
            if k in row:
                grand_y[k] = grand_y.get(k, 0) + row[k]
    # tỷ lệ tổng: tính lại từ 2 khoá đầu (late/urgent ÷ trans/orders) nếu có
    _fill_rates(grand_m, grand_y, subcols, year_cols)
    write_row(DATA0, "TỔNG CỘNG", "", lambda mk: grand_m.get(mk), grand_y, total=True)

    # ----- Các dòng đối tượng (tô đỏ khi tỷ lệ năm > 30%) -----
    r = DATA0 + 1
    for i, row in enumerate(rows, 1):
        warn = bool(warn_col_key) and (row.get(warn_col_key, 0) or 0) > 30
        write_row(r, row["key"], i, lambda mk, row=row: row.get("m", {}).get(mk), row, warn=warn)
        r += 1

    _finalize(ws, entity_label, nmonth, nsub, len(year_cols), col_month0)


def _fill_rates(grand_m, grand_y, subcols, year_cols):
    """Tính lại cột % cho dòng TỔNG CỘNG = (khoá phần) / (khoá tổng) * 100.
    Quy ước: cột % dựa trên 2 cột số đứng trước nó (phần, tổng) theo thứ tự form."""
    def rate_of(obj):
        # tìm cặp (tổng, phần): cột số đầu = tổng giao dịch/đặt, cột số 2 = trễ/gấp
        nums = [k for _h, k, f in subcols if f != S.FMT_PCT]
        if len(nums) >= 2:
            whole, part = obj.get(nums[0], 0), obj.get(nums[1], 0)
            return round(part / whole * 100, 2) if whole else 0
        return 0
    pct_sub = [k for _h, k, f in subcols if f == S.FMT_PCT]
    for mk, obj in grand_m.items():
        for k in pct_sub:
            obj[k] = rate_of(obj)
    ynums = [k for _h, k, f in year_cols if f != S.FMT_PCT]
    for _h, k, f in year_cols:
        if f == S.FMT_PCT and len(ynums) >= 2:
            whole, part = grand_y.get(ynums[0], 0), grand_y.get(ynums[1], 0)
            grand_y[k] = round(part / whole * 100, 2) if whole else 0


def _finalize(ws, entity_label, nmonth, nsub, n_year, col_month0):
    ws.freeze_panes = ws.cell(row=8, column=col_month0)
    ws.column_dimensions["A"].width = 5
    ws.column_dimensions["B"].width = max(18, len(entity_label) + 4)
    last = col_month0 + nmonth * nsub + n_year - 1
    for c in range(col_month0, last + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    for r in (6, 7):
        ws.row_dimensions[r].height = 30
