"""Form 16 — BÁO CÁO CHI PHÍ VẬN CHUYỂN: 2 bảng (tóm tắt theo chành + chi tiết theo đơn)."""
from openpyxl.utils import get_column_letter

from . import styles as S

_SUM_HEAD = ["ĐƠN VỊ VẬN CHUYỂN", "THÁNG", "TẦN SUẤT PHÁT SINH",
             "TỔNG GIÁ TRỊ ĐƠN HÀNG", "TỔNG CHI PHÍ VẬN CHUYỂN", "TỶ LỆ"]
_DETAIL_HEAD = ["Đơn vị vận chuyển", "Tháng", "Mã VTBB/NL", "Mã đơn Misa", "Số hóa đơn",
                "Ngày nhận thực tế", "Số lượng đặt NCC", "Số lượng đã nhận",
                "Thành tiền đơn hàng", "Thành tiền vận chuyển", "TỶ LỆ"]


def _hrow(ws, r, headers):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=r, column=i, value=h)
        S.style_header_cell(c)


def _num(ws, r, c, v, fmt, *, warn=False):
    cell = ws.cell(row=r, column=c, value=v)
    S.style_body_cell(cell, numeric=True, warn=warn)
    if v is not None:
        cell.number_format = fmt
    return cell


def _txt(ws, r, c, v):
    cell = ws.cell(row=r, column=c, value=v)
    S.style_body_cell(cell, numeric=False)
    return cell


def build_shipping_sheet(ws, shipping, shipping_detail):
    ws.cell(row=1, column=1, value=(
        "Báo cáo tần suất chành, chi phí chành, tỷ lệ chi phí chành trên đơn hàng")).font = S.FONT_DESC
    ws.cell(row=2, column=1, value="BÁO CÁO CHI PHÍ VẬN CHUYỂN").font = S.FONT_TITLE

    # ===== Bảng 1: tóm tắt theo (chành × tháng) =====
    r = 4
    _hrow(ws, r, _SUM_HEAD)
    r += 1
    for row in sorted(shipping, key=lambda x: -x.get("ship_cost", 0)):
        for mk, c in sorted(row.get("m", {}).items()):
            _txt(ws, r, 1, row["key"])
            _txt(ws, r, 2, f"{mk[5:7]}/{mk[:4]}")
            _num(ws, r, 3, c.get("freq", 0), S.FMT_INT)
            _num(ws, r, 4, c.get("order_value", 0), S.FMT_INT)
            _num(ws, r, 5, c.get("ship_cost", 0), S.FMT_INT)
            _num(ws, r, 6, (c.get("rate", 0) or 0) / 100.0, S.FMT_PCT)
            r += 1

    # ===== Bảng 2: chi tiết theo đơn =====
    r += 2
    _hrow(ws, r, _DETAIL_HEAD)
    r += 1
    for d in shipping_detail:
        _txt(ws, r, 1, d.get("carrier", ""))
        _txt(ws, r, 2, d.get("month", ""))
        _txt(ws, r, 3, d.get("product_code", ""))
        _txt(ws, r, 4, d.get("misa_code", ""))
        _txt(ws, r, 5, d.get("invoice_no", ""))
        _txt(ws, r, 6, d.get("received_date", ""))
        _num(ws, r, 7, d.get("qty_order", 0), S.FMT_INT)
        _num(ws, r, 8, d.get("qty_received", 0), S.FMT_INT)
        _num(ws, r, 9, d.get("order_amount", 0), S.FMT_INT)
        _num(ws, r, 10, d.get("ship_amount", 0), S.FMT_INT)
        _num(ws, r, 11, (d.get("rate", 0) or 0) / 100.0, S.FMT_PCT)
        r += 1

    widths = [22, 10, 16, 14, 14, 14, 14, 14, 18, 18, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
