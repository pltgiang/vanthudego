"""Điều phối dựng workbook Excel báo cáo mua hàng — khớp form thumua1 sheet 12–16.

`build_report_workbook(data, sheet)` với data = report_service.compute(..., full_detail=True).
sheet ∈ nspt | item_group | supplier | department | shipping | all."""
from io import BytesIO

from openpyxl import Workbook

from . import styles as S
from .pivot import build_pivot_sheet
from .sheet_shipping import build_shipping_sheet

INT, PCT = S.FMT_INT, S.FMT_PCT

# Đặc tả 4 form ma trận: (data_key, tab_name, spec cho build_pivot_sheet)
PIVOT_SPECS = {
    "nspt": {
        "tab": "12. BÁO CÁO NSPT",
        "desc": ("Báo cáo nhân sự phụ trách làm bao nhiêu đơn mua và số đơn không "
                 "theo kịp quy định trong tháng/năm"),
        "title": "BÁO CÁO NHÂN SỰ PHỤ TRÁCH",
        "entity_label": "NSPT",
        "subcols": [("Tổng số đơn", "orders", INT), ("Tổng đơn bị trễ", "late", INT),
                    ("Tổng đơn đúng hẹn", "ontime", INT), ("Tổng đơn giao sớm", "early", INT),
                    ("Tỷ lệ trễ đơn", "rate", PCT)],
        "year_cols": [("TỔNG SỐ LẦN GIAO DỊCH", "orders", INT),
                      ("TỔNG SỐ LẦN TRỄ", "late", INT), ("TỶ LỆ TRỄ ĐƠN", "rate", PCT)],
        "warn_col_key": "rate",
    },
    "item_group": {
        "tab": "13. BÁO CÁO PHÂN LOẠI VTBBNL",
        "desc": "Báo cáo tần suất mua và chi phí mua loại VTBB/NL trong tháng/năm",
        "title": "BÁO CÁO THỐNG KÊ TẦN SUẤT MUA CỦA CÁC LOẠI VTBB/NL",
        "entity_label": "LOẠI VTBB/NL",
        "subcols": [("Tổng số lần giao dịch", "trans", INT), ("Tổng chi phí mua", "cost", INT)],
        "year_cols": [("TỔNG SỐ LẦN GIAO DỊCH", "trans", INT), ("TỔNG CP MUA", "cost", INT)],
        "warn_col_key": None,
    },
    "supplier": {
        "tab": "14. BÁO CÁO NCC",
        "desc": ("Báo cáo NCC thể hiện tần suất phát sinh, số lần trễ tiến độ, "
                 "tỷ lệ trễ theo tháng/năm"),
        "title": "BÁO CÁO THỐNG KÊ CÁC LẦN GIAO DỊCH NCC",
        "entity_label": "NCC",
        "subcols": [("Tổng số lần giao dịch", "trans", INT), ("Số lần trễ", "late", INT),
                    ("Tỷ lệ", "rate", PCT)],
        "year_cols": [("TỔNG SỐ LẦN GIAO DỊCH", "trans", INT),
                      ("TỔNG SỐ LẦN TRỄ", "late", INT), ("TỶ LỆ TRỄ", "rate", PCT)],
        "warn_col_key": "rate",
    },
    "department": {
        "tab": "15. BÁO CÁO BỘ PHẬN",
        "desc": ("Báo cáo tổng số đơn hàng BP đặt, số đơn trong diện gấp và tỷ lệ gấp"),
        "title": "BÁO CÁO THỐNG KÊ ĐẶT HÀNG THEO BỘ PHẬN",
        "entity_label": "BỘ PHẬN",
        "subcols": [("Tổng số lần đặt hàng", "orders", INT), ("Số lần gấp", "urgent", INT),
                    ("Tỷ lệ", "rate", PCT)],
        "year_cols": [("TỔNG SỐ LẦN ĐẶT HÀNG", "orders", INT),
                      ("TỔNG SỐ LẦN YÊU CẦU GẤP", "urgent", INT), ("TỶ LỆ GẤP", "rate", PCT)],
        "warn_col_key": "rate",
    },
}

SHEET_ORDER = ["nspt", "item_group", "supplier", "department", "shipping"]


def _add_pivot(wb, key, data):
    spec = PIVOT_SPECS[key]
    ws = wb.create_sheet(title=spec["tab"])
    build_pivot_sheet(
        ws, desc=spec["desc"], title=spec["title"], entity_label=spec["entity_label"],
        months=data["months"], rows=data.get(key, []),
        subcols=spec["subcols"], year_cols=spec["year_cols"], warn_col_key=spec["warn_col_key"],
    )


def _add_shipping(wb, data):
    ws = wb.create_sheet(title="16. BÁO CÁO CHI PHÍ VẬN CHUYỂN")
    build_shipping_sheet(ws, data.get("shipping", []), data.get("shipping_detail", []))


def build_report_workbook(data: dict, sheet: str) -> BytesIO:
    wb = Workbook()
    wb.remove(wb.active)   # bỏ sheet mặc định

    keys = SHEET_ORDER if sheet == "all" else [sheet]
    for k in keys:
        if k == "shipping":
            _add_shipping(wb, data)
        elif k in PIVOT_SPECS:
            _add_pivot(wb, k, data)
        else:
            raise ValueError(f"sheet không hợp lệ: {sheet}")

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
