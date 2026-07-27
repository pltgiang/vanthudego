"""Style & định dạng dùng chung cho các sheet Excel báo cáo mua hàng."""
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# Số nguyên / tiền: '#,##0'  ·  Tỷ lệ %: '0.0%' (giá trị ghi vào là phân số, đã chia 100)
FMT_INT = "#,##0"
FMT_PCT = "0.0%"

FONT_TITLE = Font(bold=True, size=13)
FONT_DESC = Font(italic=True, size=10, color="555555")
FONT_HEAD = Font(bold=True, size=10, color="FFFFFF")
FONT_TOTAL = Font(bold=True, size=10)
FONT_WARN_NOTE = Font(bold=True, size=10, color="C00000")

FILL_HEAD = PatternFill("solid", fgColor="305496")      # xanh đậm header
FILL_TOTAL = PatternFill("solid", fgColor="DDEBF7")      # xanh nhạt dòng tổng cộng
FILL_WARN = PatternFill("solid", fgColor="FFC7CE")       # đỏ nhạt cảnh báo >30%

_side = Side(style="thin", color="B0B0B0")
BORDER = Border(left=_side, right=_side, top=_side, bottom=_side)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def style_header_cell(cell):
    cell.font = FONT_HEAD
    cell.fill = FILL_HEAD
    cell.alignment = CENTER
    cell.border = BORDER


def style_body_cell(cell, *, numeric=True, total=False, warn=False):
    cell.border = BORDER
    cell.alignment = RIGHT if numeric else LEFT
    if total:
        cell.font = FONT_TOTAL
        cell.fill = FILL_TOTAL
    if warn:
        cell.fill = FILL_WARN
