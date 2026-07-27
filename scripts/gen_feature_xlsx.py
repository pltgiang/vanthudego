# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import Counter

wb = openpyxl.Workbook()

NAVY = "1B2559"; TEAL = "00AEEF"; GRAY = "64748B"
HFILL = PatternFill("solid", fgColor=NAVY)
GFILL = PatternFill("solid", fgColor="EEF2F6")
OK = PatternFill("solid", fgColor="DCFCE7"); OKF = Font(color="15803D", bold=True)
WIP = PatternFill("solid", fgColor="FEF3C7"); WIPF = Font(color="B45309", bold=True)
thin = Side(style="thin", color="D9E0EA")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
HFONT = Font(color="FFFFFF", bold=True, size=11)
WRAP = Alignment(wrap_text=True, vertical="top")
CEN = Alignment(horizontal="center", vertical="center")


def style_header(ws, row, ncol):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HFILL; cell.font = HFONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BORDER


# ---------- SHEET: DANH SÁCH TÍNH NĂNG ----------
DONE = "Đã xong"; WIPS = "Đang hoàn thiện"
feats = [
    ("Mua hàng", "Trang chủ / Dashboard", "Thống kê tổng quan, duyệt nhanh PYC/đơn hàng", "/ (Trang chủ)", 1, DONE, ""),
    ("Mua hàng", "Yêu cầu mua hàng (PYC)", "Tạo/duyệt/từ chối, in A4, xóa nhiều", "Danh sách + Chi tiết + In", 3, DONE, ""),
    ("Mua hàng", "Yêu cầu khảo sát", "Tạo/duyệt + màn Xử lý (tạo phương án, gán NSTM)", "DS + Chi tiết + Xử lý", 3, WIPS, "Còn giai đoạn 5C/5D"),
    ("Mua hàng", "Phiếu khảo sát", "2 bảng NCC + Sản phẩm, duyệt từng dòng, đính kèm", "DS + Chi tiết (NCC/SP)", 4, WIPS, "Đang tinh chỉnh UI/nghiệp vụ"),
    ("Mua hàng", "Đơn mua hàng (PO)", "Items + giao hàng trong popup, in A4 ngang", "DS + Chi tiết + In (2 mẫu)", 4, DONE, ""),
    ("Mua hàng", "Báo cáo mua hàng", "Báo cáo/thống kê mua hàng", "/reports", 1, DONE, ""),
    ("Mua hàng", "Báo cáo khảo sát", "Báo cáo kết quả khảo sát", "/survey-report", 1, DONE, ""),
    ("Kho & Công nợ", "Tồn kho", "Xem tồn + điều chỉnh", "/inventory", 1, DONE, ""),
    ("Kho & Công nợ", "Nhận hàng (ghi tồn)", "Ghi tồn tự động (ngầm) khi nhận hàng theo PO", "(ngầm, không màn riêng)", 0, DONE, ""),
    ("Kho & Công nợ", "Công nợ", "Sinh tự động, xem tuổi nợ", "/payables", 1, DONE, ""),
    ("Kho & Công nợ", "Yêu cầu thanh toán", "1 NCC/phiếu, gom nhiều PO, in được", "DS + Chi tiết + In", 3, DONE, ""),
    ("Danh mục", "Nhà cung cấp", "Quản lý NCC", "DS + Chi tiết", 2, DONE, ""),
    ("Danh mục", "Sản phẩm", "Danh mục sản phẩm (~6.760)", "DS + Chi tiết", 2, DONE, "Import 1 chiều từ Google Sheet"),
    ("Danh mục", "Hợp đồng", "Quản lý hợp đồng NCC", "DS + Chi tiết", 2, DONE, ""),
    ("Danh mục", "Kho", "Danh mục kho", "CRUD", 1, DONE, ""),
    ("Danh mục", "Đơn vị tính", "Danh mục ĐVT", "CRUD", 1, DONE, ""),
    ("Danh mục", "Phân loại (nhóm hàng)", "Danh mục phân loại", "CRUD", 1, DONE, ""),
    ("Danh mục", "Thương hiệu", "Danh mục thương hiệu", "CRUD", 1, DONE, ""),
    ("Danh mục", "Phòng ban", "Danh mục phòng ban", "CRUD", 1, DONE, ""),
    ("Danh mục", "Phân công phụ trách", "Gán người phụ trách theo phân loại", "DS + Thêm", 2, DONE, ""),
    ("Danh mục", "Công ty", "Danh mục công ty", "CRUD", 1, DONE, ""),
    ("Hệ thống", "Nhân sự", "Danh mục nhân viên", "CRUD", 1, DONE, ""),
    ("Hệ thống", "Người dùng", "Tài khoản đăng nhập", "DS + Chi tiết", 2, DONE, ""),
    ("Hệ thống", "Vai trò & Phân quyền (RBAC)", "Phân quyền chi tiết theo entity/action + phạm vi", "/roles", 1, DONE, ""),
    ("Hệ thống", "Cấu hình hệ thống", "Thông số hệ thống (SMTP, R2, ...)", "/settings", 1, DONE, ""),
    ("Nền tảng", "Xác thực & Đăng nhập", "Đăng nhập, quên/đặt lại mật khẩu (JWT)", "Login/Quên/Đặt lại MK", 3, DONE, ""),
    ("Nền tảng", "Hệ thống thông báo", "Chuông thông báo theo sự kiện", "(dropdown chuông)", 0, DONE, "VPS tắt email, chỉ chuông"),
    ("Nền tảng", "Đính kèm file", "Upload R2 (fallback local), 2 bảng file/liên kết", "(trong các phiếu)", 0, DONE, ""),
    ("Nền tảng", "Nhật ký thao tác (Audit)", "Ghi lịch sử thao tác trên phiếu", "(trong màn chi tiết)", 0, DONE, ""),
    ("Nền tảng", "Toast & Modal xác nhận", "Thông báo + xác nhận/nhập lý do dùng chung", "(toàn hệ thống)", 0, DONE, ""),
]

ws = wb.active; ws.title = "Danh sách tính năng"
cols = ["STT", "Nhóm", "Tính năng / Chức năng", "Mô tả", "Màn hình chính", "Số MH", "Hiện trạng", "Ghi chú"]
ws.append(cols); style_header(ws, 1, len(cols))
for i, (grp, name, desc, scr, nmh, st, note) in enumerate(feats, 1):
    ws.append([i, grp, name, desc, scr, nmh, st, note])
    r = ws.max_row
    for c in range(1, len(cols) + 1):
        cell = ws.cell(row=r, column=c); cell.border = BORDER; cell.alignment = WRAP
    ws.cell(row=r, column=1).alignment = CEN
    ws.cell(row=r, column=6).alignment = CEN
    stc = ws.cell(row=r, column=7); stc.alignment = CEN
    if st == DONE:
        stc.fill = OK; stc.font = OKF
    else:
        stc.fill = WIP; stc.font = WIPF
for i, w in enumerate([6, 15, 30, 40, 26, 8, 18, 26], 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.freeze_panes = "A2"; ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

# ---------- SHEET: DANH SÁCH MÀN HÌNH ----------
screens = [
    ("Đăng nhập", "/login", "Trang", "Nền tảng"),
    ("Quên mật khẩu", "/forgot-password", "Trang", "Nền tảng"),
    ("Đặt lại mật khẩu", "/reset-password", "Trang", "Nền tảng"),
    ("Trang chủ / Dashboard", "/", "Dashboard", "Mua hàng"),
    ("Yêu cầu mua hàng — Danh sách", "/purchase-requests", "Danh sách", "Mua hàng"),
    ("Yêu cầu mua hàng — Chi tiết", "/purchase-requests/:id", "Chi tiết", "Mua hàng"),
    ("Yêu cầu khảo sát — Danh sách", "/survey-requests", "Danh sách", "Mua hàng"),
    ("Yêu cầu khảo sát — Chi tiết", "/survey-requests/:id", "Chi tiết", "Mua hàng"),
    ("Yêu cầu khảo sát — Xử lý", "/survey-requests/:id/process", "Xử lý", "Mua hàng"),
    ("Phiếu khảo sát — Danh sách", "/surveys", "Danh sách", "Mua hàng"),
    ("Phiếu khảo sát — Chi tiết", "/surveys/:id", "Chi tiết", "Mua hàng"),
    ("Tạo KS Nhà cung cấp", "/surveys-supplier/:id", "Chi tiết", "Mua hàng"),
    ("Tạo KS Sản phẩm", "/surveys-product/:id", "Chi tiết", "Mua hàng"),
    ("Đơn mua hàng — Danh sách", "/purchase-orders", "Danh sách", "Mua hàng"),
    ("Đơn mua hàng — Chi tiết", "/purchase-orders/:id", "Chi tiết", "Mua hàng"),
    ("Yêu cầu thanh toán — Danh sách", "/payment-requests", "Danh sách", "Kho & Công nợ"),
    ("Yêu cầu thanh toán — Chi tiết", "/payment-requests/:id", "Chi tiết", "Kho & Công nợ"),
    ("Tồn kho", "/inventory", "Danh sách", "Kho & Công nợ"),
    ("Công nợ", "/payables", "Danh sách", "Kho & Công nợ"),
    ("Báo cáo mua hàng", "/reports", "Báo cáo", "Mua hàng"),
    ("Báo cáo khảo sát", "/survey-report", "Báo cáo", "Mua hàng"),
    ("Nhà cung cấp", "/suppliers (+/:id)", "DS + Chi tiết", "Danh mục"),
    ("Hợp đồng", "/contracts (+/:id)", "DS + Chi tiết", "Danh mục"),
    ("Phân công phụ trách", "/category-assignees (+/new)", "DS + Thêm", "Danh mục"),
    ("Sản phẩm / Kho / ĐVT / Phân loại / Thương hiệu / Công ty / Nhân sự", "/:entity (+/:id)", "CRUD chung", "Danh mục / Hệ thống"),
    ("Người dùng", "/users (+/:id)", "DS + Chi tiết", "Hệ thống"),
    ("Vai trò & Phân quyền", "/roles", "Cấu hình", "Hệ thống"),
    ("Cấu hình hệ thống", "/settings", "Cấu hình", "Hệ thống"),
    ("In — Yêu cầu mua hàng", "/print/purchase-request/:id", "In A4", "Mua hàng"),
    ("In — Đơn mua hàng", "/print/purchase-order/:id", "In A4", "Mua hàng"),
    ("In — Đơn mua hàng (mẫu MH)", "/print/purchase-order-mh/:id", "In A4", "Mua hàng"),
    ("In — Yêu cầu thanh toán", "/print/payment-request/:id", "In A4", "Kho & Công nợ"),
]
ws3 = wb.create_sheet("Danh sách màn hình")
c3 = ["STT", "Màn hình", "Đường dẫn (route)", "Loại", "Nhóm"]
ws3.append(c3); style_header(ws3, 1, len(c3))
for i, (name, route, typ, grp) in enumerate(screens, 1):
    ws3.append([i, name, route, typ, grp]); r = ws3.max_row
    for c in range(1, len(c3) + 1):
        cell = ws3.cell(row=r, column=c); cell.border = BORDER; cell.alignment = WRAP
    ws3.cell(row=r, column=1).alignment = CEN
for i, w in enumerate([6, 50, 32, 16, 20], 1):
    ws3.column_dimensions[get_column_letter(i)].width = w
ws3.freeze_panes = "A2"; ws3.auto_filter.ref = f"A1:{get_column_letter(len(c3))}1"

# ---------- SHEET: TỔNG QUAN ----------
ws1 = wb.create_sheet("Tổng quan", 0)
n_feat = len(feats)
n_done = sum(1 for f in feats if f[5] == DONE)
n_wip = n_feat - n_done
n_screen_est = sum(f[4] for f in feats)
n_screen_rows = len(screens)
ws1["B2"] = "MINI TOOL QUẢN LÝ THU MUA - DEGO HOLDING"; ws1["B2"].font = Font(bold=True, size=16, color=NAVY)
ws1["B3"] = "Danh sách Tính năng & Hiện trạng"; ws1["B3"].font = Font(size=12, color=TEAL, bold=True)
ws1["B4"] = "Cập nhật: 09/07/2026"; ws1["B4"].font = Font(size=10, color=GRAY, italic=True)
kpis = [
    ("Tổng số tính năng/chức năng", n_feat), ("- Đã xong", n_done), ("- Đang hoàn thiện", n_wip),
    ("Tỉ lệ hoàn thành", f"{round(n_done / n_feat * 100)}%"), ("Số nhóm chức năng", 5),
    ("Số màn hình (ước tính)", n_screen_est), ("Số route/màn liệt kê", n_screen_rows),
    ("Số module backend", 29),
]
r = 6
for label, val in kpis:
    lc = ws1.cell(row=r, column=2, value=label); lc.font = Font(bold=True, color=NAVY, size=11); lc.fill = GFILL; lc.border = BORDER
    vc = ws1.cell(row=r, column=3, value=val); vc.font = Font(bold=True, size=12, color=TEAL); vc.alignment = CEN; vc.border = BORDER
    r += 1
r += 1
ws1.cell(row=r, column=2, value="Theo nhóm").font = Font(bold=True, color=NAVY)
r += 1
h1 = ws1.cell(row=r, column=2, value="Nhóm"); h1.font = HFONT; h1.fill = HFILL; h1.border = BORDER
h2 = ws1.cell(row=r, column=3, value="Số tính năng"); h2.font = HFONT; h2.fill = HFILL; h2.alignment = CEN; h2.border = BORDER
cnt = Counter(f[0] for f in feats)
for grp in ["Mua hàng", "Kho & Công nợ", "Danh mục", "Hệ thống", "Nền tảng"]:
    r += 1
    ws1.cell(row=r, column=2, value=grp).border = BORDER
    vc = ws1.cell(row=r, column=3, value=cnt[grp]); vc.alignment = CEN; vc.border = BORDER
ws1.column_dimensions["A"].width = 3
ws1.column_dimensions["B"].width = 34
ws1.column_dimensions["C"].width = 18
ws1.sheet_view.showGridLines = False

out = "doc/danh-sach-tinh-nang.xlsx"
wb.save(out)
print("Da tao:", out)
print("Tinh nang:", n_feat, "| Da xong:", n_done, "| Dang hoan thien:", n_wip,
      "| Man hinh uoc tinh:", n_screen_est, "| Route liet ke:", n_screen_rows)
